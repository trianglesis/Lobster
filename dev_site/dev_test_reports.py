import datetime
import openpyxl
from django.db.utils import IntegrityError
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Fill

from django.db.models.functions import TruncMonth, TruncDay

from django.db.models import Q
from django.db.models.aggregates import Sum, Avg, Max, Min, StdDev, Variance, Count

if __name__ == "__main__":

    import django
    import logging

    log = logging.getLogger("octo.octologger")
    django.setup()
    from django.conf import settings

    from octo_tku_patterns.model_views import TestReportsView
    from octo_tku_patterns.models import TestReports


    def get_stats(**kwargs):
        # print(f"Selecting by: {kwargs}")
        report_date_time = kwargs.get('report_date_time')
        queryset = TestReports.objects.all()
        if kwargs.get("test_type"):
            queryset = queryset.filter(test_type__exact=kwargs.get("test_type"))
        else:
            queryset = queryset.filter(tkn_branch__isnull=False)
        if kwargs.get('addm_name'):
            queryset = queryset.filter(addm_name__exact=kwargs.get('addm_name'))
        if kwargs.get('tkn_branch'):
            queryset = queryset.filter(tkn_branch__exact=kwargs.get('tkn_branch'))
        if kwargs.get('pattern_library'):
            queryset = queryset.filter(pattern_library__exact=kwargs.get('pattern_library'))

        if kwargs.get('report_date_time'):
            if not isinstance(report_date_time, datetime.date):
                print(
                    f"Warning: report_date_time should bw a datetime object, if not - will use today date by default!")
                date_var = datetime.date.today()
            else:
                date_var = report_date_time
            if kwargs.get('date_t') == 'year':
                queryset = queryset.filter(Q(report_date_time__year=date_var.year))
            elif kwargs.get('date_t') == 'month':
                queryset = queryset.filter(
                    Q(report_date_time__year=date_var.year, report_date_time__month=date_var.month))
            elif kwargs.get('date_t') == 'days_range':
                prev = date_var - datetime.timedelta(days=kwargs.get('days_range'))
                tomorrow = date_var + datetime.timedelta(days=1)
                print(f'Select betweeen: {prev} : {tomorrow}')
                queryset = queryset.filter(Q(report_date_time__range=[prev, tomorrow]))
            else:
                queryset = queryset.filter(
                    Q(report_date_time__year=date_var.year, report_date_time__month=date_var.month,
                      report_date_time__day=date_var.day))
        return queryset


    def insert_digest():
        tests = TestReportsView.objects.all().values(
            # 'test_type',
            # 'tkn_branch',
            # 'pattern_library',
            # 'addm_name',
            # 'addm_v_int',
            # 'tests_count',
            # 'patterns_count',
            # 'fails',
            # 'error',
            # 'passed',
            # 'skipped',
        )

        for item in tests:
            try:
                save_tst = TestReports(**item)
                save_tst.save(force_insert=True)
            except IntegrityError as e:
                print(f'Duplicate: {e}')


    def select_stats(**kwargs):
        by_value = kwargs.get('grouping')
        queryset = get_stats(**kwargs)
        queryset = queryset.values(by_value).order_by(by_value).annotate(
            date=Max('report_date_time'),
            addm_v_int=Max('addm_v_int'),
            tests_sum=Sum('tests_count'),
            patterns_sum=Sum('patterns_count'),
            errors_sum=Sum('error'),
            fails_sum=Sum('fails'),
            passed_sum=Sum('passed'),
            skipped_sum=Sum('skipped'),
        )
        # print(f"{queryset.query}")
        return queryset


    def select_stats_daily(**kwargs):
        queryset = get_stats(**kwargs)
        queryset = queryset.annotate(day=TruncDay('report_date_time')).values('day').order_by('day').annotate(
            date=Max('report_date_time'),
            addm_v_int=Max('addm_v_int'),
            tests_sum=Sum('tests_count'),
            patterns_sum=Sum('patterns_count'),
            errors_sum=Sum('error'),
            fails_sum=Sum('fails'),
            passed_sum=Sum('passed'),
            skipped_sum=Sum('skipped'),
        )
        # print(f"{queryset.query}")
        return queryset


    def make_queries(day_range=30):
        for day in range(day_range):
            date_var_ = datetime.datetime.now() - datetime.timedelta(days=day)
            # date_var_ = datetime.date.today()
            queryset_ = select_stats(
                tkn_branch='tkn_ship',
                test_type='tku_patterns',
                grouping='tkn_branch',
                report_date_time=date_var_,
                # addm_name='gargoyle',
                # pattern_library='CORE',
            )
            for row in queryset_:
                print(f"Result for day {date_var_.strftime('%Y-%m-%d')}: {row}")


    def make_query(day, grouping='tkn_branch', tkn_branch='tkn_ship', test_type='tku_patterns', addm_name=None,
                   pattern_library=None):
        date_var_ = datetime.datetime.now()
        date_var_ = date_var_ - datetime.timedelta(days=day)
        # print(f"Query stats for: {date_var_.strftime('%m/%d')}")
        sel_kw = dict(
            tkn_branch=tkn_branch,
            test_type=test_type,
            grouping=grouping,
            report_date_time=date_var_,
            # pattern_library='CORE',
        )
        if addm_name:
            sel_kw.update(addm_name=addm_name)
        if pattern_library:
            sel_kw.update(pattern_library=pattern_library)

        queryset_ = select_stats(**sel_kw)
        if queryset_:
            return queryset_[0]
        else:
            return dict(
                date=date_var_,
                tests_sum=0,
            )


    def make_q_days_range(days=31, grouping='tkn_branch', tkn_branch='tkn_ship', test_type='tku_patterns',
                          addm_name=None, pattern_library=None):
        """
        Select records from today to previous 31 days date.

        :param days:
        :param grouping:
        :param tkn_branch:
        :param test_type:
        :param addm_name:
        :param pattern_library:
        :return:
        """
        now = datetime.datetime.now(tz=timezone.utc).replace(hour=00, minute=00, second=00)
        queryset = get_stats(
            tkn_branch=tkn_branch,
            test_type=test_type,
            grouping=grouping,
            addm_name=addm_name,
            pattern_library=pattern_library,
            report_date_time=now,
            date_t='days_range',
            days_range=days,
        )
        queryset = queryset.annotate(day=TruncDay('report_date_time')).values('day').order_by('day').annotate(
            date=Max('report_date_time'),
            addm_v_int=Max('addm_v_int'),
            tests_sum=Sum('tests_count'),
            patterns_sum=Sum('patterns_count'),
            errors_sum=Sum('error'),
            fails_sum=Sum('fails'),
            passed_sum=Sum('passed'),
            skipped_sum=Sum('skipped'),
        )
        return queryset


    def make_q_month(grouping='tkn_branch', tkn_branch='tkn_ship', test_type='tku_patterns', addm_name=None,
                     pattern_library=None):
        """
        Select current month records from 1st day.

        :param grouping:
        :param tkn_branch:
        :param test_type:
        :param addm_name:
        :param pattern_library:
        :return:
        """
        today_y = datetime.datetime.now(tz=timezone.utc).replace(hour=00, minute=00, second=00)
        start_date = datetime.datetime(today_y.year, today_y.month, 1)
        queryset = get_stats(
            tkn_branch=tkn_branch,
            test_type=test_type,
            grouping=grouping,
            addm_name=addm_name,
            pattern_library=pattern_library,
            report_date_time=start_date,
            date_t='month',
        )
        queryset = queryset.annotate(day=TruncDay('report_date_time')).values('day').order_by('day').annotate(
            date=Max('report_date_time'),
            addm_v_int=Max('addm_v_int'),
            tests_sum=Sum('tests_count'),
            patterns_sum=Sum('patterns_count'),
            errors_sum=Sum('error'),
            fails_sum=Sum('fails'),
            passed_sum=Sum('passed'),
            skipped_sum=Sum('skipped'),
        )
        return queryset


    def make_q_year(grouping='tkn_branch', tkn_branch='tkn_ship', test_type='tku_patterns', addm_name=None,
                    pattern_library=None):
        """
        Select current month records from 1st day of the year.

        :param grouping:
        :param tkn_branch:
        :param test_type:
        :param addm_name:
        :param pattern_library:
        :return:
        """
        today_y = datetime.datetime.now(tz=timezone.utc).replace(hour=00, minute=00, second=00)
        start_date = datetime.datetime(today_y.year, 1, 1)
        queryset = get_stats(
            tkn_branch=tkn_branch,
            test_type=test_type,
            grouping=grouping,
            addm_name=addm_name,
            pattern_library=pattern_library,
            report_date_time=start_date,
            date_t='year',
        )
        queryset = queryset.annotate(day=TruncDay('report_date_time')).values('day').order_by('day').annotate(
            date=Max('report_date_time'),
            addm_v_int=Max('addm_v_int'),
            tests_sum=Sum('tests_count'),
            patterns_sum=Sum('patterns_count'),
            errors_sum=Sum('error'),
            fails_sum=Sum('fails'),
            passed_sum=Sum('passed'),
            skipped_sum=Sum('skipped'),
        )
        return queryset


    def workbook_create(filename=None):
        if not filename:
            filename = 'test.xlsx'
        else:
            filename = f'{filename}.xlsx'

        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Last_30_days"
        ws1.page_setup.fitToWidth = 1

        font_B = Font(name='Calibri', size=9, bold=True)
        font_N = Font(name='Calibri', size=9, bold=False)
        border_t = Border(
            left=Side(border_style='thin', color='FF000000'),
            right=Side(border_style='thin', color='FF000000'),
            top=Side(border_style='thin', color='FF000000'),
            bottom=Side(border_style='thin', color='FF000000'),
        )
        border_s = Border(
            left=Side(border_style='thin', color='FF000000'),
            right=Side(border_style='thin', color='FF000000'),
            # top=Side(border_style='thin', color='FF000000'),
            # bottom=Side(border_style='thin', color='FF000000'),
        )
        alignment_c = Alignment(
            horizontal='center',
            vertical='center',
            shrinkToFit=True)

        fill_yellow = PatternFill("solid", fgColor="00FFFF99")

        # Merge Cells:
        # ws1.merge_cells('B1:AI1')
        ws1.column_dimensions['A'].width = 2
        ws1.column_dimensions['AG'].width = 2
        ws1.column_dimensions['AJ'].width = 2
        ws1.column_dimensions['B'].width = 20
        ws1.column_dimensions['AH'].width = 20
        # HEAD
        ws1.merge_cells(start_row=1, start_column=2, end_row=2, end_column=36)
        # HEADER
        ws1["B1"] = 'Automation Status for BMC Discovery Content'
        ws1["B1"].font = font_B
        # MONTH VIEW
        ws1["B5"] = 'Pass %'
        ws1["B6"] = '# of Failures'
        ws1["B7"] = '# of Tests Executed'
        ws1["B8"] = '# of Tests Not  Executed'
        # MONTH STATS
        ws1["AH3"] = 'Last 30 days stats:'
        ws1["AH3"].font = font_B
        ws1["AH5"] = '# of Days at 100%:'
        ws1["AH6"] = 'Avg Pass Rate:'
        ws1["AH7"] = 'Avg # of Failures:'

        """
        C4:AF4 - Date %m-%d
        C5:AF4 - Pass %
        C6:AF6 - Fails #
        C7:AF7 - Run test #
        C8:AF8 - Not run test #
        
        AI5 - Days 100% Success
        AI6 - Avg Pass Rate
        AI7 - Avg # of Failures
        """

        # Get ROWs 4:8
        for _r in range(4, 9):
            # Get COLs A:AI
            for _c in range(1, 36):
                # Get COLs: B:AF
                cell = ws1.cell(column=_c, row=_r)
                # print(f"COL: {cell.column_letter}:{_r}")
                # Make borders B:AF
                if _c in range(2, 33):
                    # Get COLs C:AF and not on rows 4:8
                    if _c in range(3, 33) and not _r in [4, 8]:
                        cell.border = border_s
                        cell.alignment = alignment_c
                    # Get COLs C:AF on rows 4:8
                    else:
                        cell.border = border_t
                    # Make COL thinner C:AF
                    if _c in range(3, 33) and _r == 4:
                        ws1.column_dimensions[cell.column_letter].width = 6
                # Make font
                # Get COLs A:AJ
                if _c in range(1, 36):
                    cell.font = font_N
                # Make summary table borders
                # Get ROWs 4:7
                if _r in range(4, 8):
                    # Get COLs AH:AI
                    if _c in range(34, 36):
                        cell.border = border_t
        """
        Possible way to get all cells previously filled with style
         and iter COLs from each 4th to the last 8th with borders.
         While iteration over each COL - we also get its child CELLs from above C1,C2,...C8
         We only need to fill CELLs with data starting from 4th cell from above staying in current COL.
         Next COL D and all following CELLs will be filled accordingly.
         
         Iterate all COLS from ROW 4th - which is the ROW where is no merged cells!
         
         Firstly the list of COLs to iterate is reversed, so we start from the end - "AJ" col, 
          which is latest one with the style applied.
         Then we step back for 4 cols, to get as first actual COL "AF" as starting point to fill with Data.
         During iteration - increment the i variable which is used as day timedelta from today, to minus 31 day at past.
         Each incrementation calls a query for that day and fill table with a test stat data.
         
          
        """
        i = 0
        # Iter each COL in row 4
        reversed_col_l = list(ws1[4])[::-1]
        print(f"Col 4 list reversed {reversed_col_l[4:]}\n")
        for cell_r in reversed_col_l[4:]:
            if i in range(0, 31):
                print(f'Query day: {i}')
                # get queryset for last 30 days, DESC?
                test_stats = make_query(day=i)

                # Get COL literal name
                col_lit = cell_r.column_letter
                # Get CELL list from column with specific literal name:
                cell_list = list(ws1[col_lit])
                # print(f"COL: {col_lit} CELL list: {cell_list}")

                if test_stats:
                    print(test_stats)
                    # Here assign required cell with some data:
                    cell_list[3].value = test_stats['date'].strftime('%m/%d')  # 'date'
                    cell_list[3].fill = fill_yellow
                    if test_stats['tests_sum']:
                        cell_list[
                            4].value = f"{round(100 * (int(test_stats['passed_sum']) + int(test_stats['skipped_sum'])) / float(test_stats['tests_sum']), 1)}%"  # 'pass %'
                        cell_list[5].value = test_stats['fails_sum']  # '# fails'
                        cell_list[6].value = test_stats['tests_sum']  # '# executed'
                        cell_list[7].value = 0  # '# not executed'

                # Get each 3rd cell, iter over last active: A4:A8
                # for cell in cell_list:
                for cell in cell_list[3:]:
                    print(f'each 3rd cell: {cell} {cell.value}')
            i += 1

        ws2 = wb.create_sheet(title="Historical")

        wb.save(filename=filename)


    def workbook_create_test(filename=None, start_row=1, start_col=2):
        if settings.DEV:
            place = ''
        else:
            place = '/home/user/TH_Octopus/UPLOAD/report/'

        if not filename:
            filename = 'test_wb.xlsx'
        else:
            filename = f'{filename}.xlsx'

        end_row = start_row + 1  # 2nd row to merge with starting
        end_column = 36  # End of table in col length - AJ col in table.

        min_col = 1  # Always start iter COLs from A col - 1st
        max_col = end_column  # Always end iter on the latest COL: 36th AJ col in table.
        min_row = start_row + 2  # Where day cols are starting to count from - 3rd after two merged
        max_row = start_row + 7  # Where day stat cols and on row  - 8th from the 1st initial

        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Last_30_days"

        font_N = Font(name='Calibri', size=9, bold=False)
        font_B = Font(name='Calibri', size=9, bold=True)

        border_t = Border(
            left=Side(border_style='thin', color='FF000000'),
            right=Side(border_style='thin', color='FF000000'),
            top=Side(border_style='thin', color='FF000000'),
            bottom=Side(border_style='thin', color='FF000000'),
        )
        border_s = Border(
            left=Side(border_style='thin', color='FF000000'),
            right=Side(border_style='thin', color='FF000000'),
            # top=Side(border_style='thin', color='FF000000'),
            # bottom=Side(border_style='thin', color='FF000000'),
        )
        alignment_c = Alignment(
            horizontal='center',
            vertical='center',
            shrinkToFit=True)

        fill_yellow = PatternFill("solid", fgColor="00FFFF99")

        # Merging cells:
        ws1.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_column)
        # HEAD
        ws1.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_column)
        # HEADER
        ws1["B1"] = 'Automation Status for BMC Discovery Content'
        ws1["B1"].font = font_B

        # HORIZONTAL Iter ROW 4:8
        rows_iter = ws1.iter_rows(min_col=min_col, max_col=max_col, min_row=min_row + 1, max_row=max_row)
        rows_iter = list(rows_iter)
        for i, _cells in enumerate(rows_iter):
            # print(f'{i} ROW, _cells: {_cells} in {i} ROW.')
            # ROW of 'date' - make formatting before.
            if i == 0:
                for c in range(2, 32):
                    ws1.column_dimensions[_cells[c].column_letter].width = 6
                    _cells[c].font = font_N
                    _cells[c].border = border_t
                    _cells[c].fill = fill_yellow
                    _cells[c].alignment = alignment_c
            # Row with data - make formatting before.
            if i in range(1, 4):
                for c in range(2, 32):
                    ws1.column_dimensions[_cells[c].column_letter].width = 6
                    _cells[c].font = font_N
                    _cells[c].border = border_s
                    _cells[c].alignment = alignment_c
            # ROW to put test report data - make formatting before.
            if i in range(0, 4):
                for c in range(2, 32):
                    ws1.column_dimensions[_cells[c].column_letter].width = 6
                    # _cells[c].value = '0'
                    _cells[c].font = font_N
                    # _cells[c].border = border_s
                    _cells[c].alignment = alignment_c
            # ROW of '#not executed' - make formatting before and fill 0 by default, as we run 100% of tests!
            if i == 4:
                for c in range(2, 32):
                    ws1.column_dimensions[_cells[c].column_letter].width = 6
                    _cells[c].value = '0'  # Do not change, we don't actually  "NOT RUN" the tests
                    _cells[c].font = font_N
                    _cells[c].border = border_t
                    _cells[c].alignment = alignment_c

        # FILL DATA HORIZONTALLY:
        # NOTE: CANNOT REVERSE!
        # for day in range(0, 30):
        #     test_stats = make_query(day=day)
        #     cell = day + 2
        #     date          = rows_iter[0][cell]
        #     pass_rate     = rows_iter[1][cell]
        #     fail_rate     = rows_iter[2][cell]
        #     exec_rate     = rows_iter[3][cell]
        #     not_exec_rate = rows_iter[4][cell]
        #
        #     print(f"Select day {day} CELL shift by {cell}")
        #
        #     # date.value = 'date'
        #     # pass_rate.value = 'pass_rate'
        #     # fail_rate.value = 'fail_rate'
        #     # exec_rate.value = 'exec_rate'
        #     # not_exec_rate.value = 'not_exec_rate'
        #
        #     date.value = test_stats['date'].strftime('%m/%d')  # 'date'
        #     if test_stats['tests_sum']:
        #         pass_rate.value = f"{round(100 * (int(test_stats['passed_sum']) + int(test_stats['skipped_sum'])) / float(test_stats['tests_sum']), 1)}%"  # 'pass %'
        #         fail_rate.value = test_stats['fails_sum']  # '# fails'
        #         exec_rate.value = test_stats['tests_sum']  # '# executed'
        #         not_exec_rate.value = 0  # '# not executed'

        # VERTICAL Iter COL A:AJ
        cols_iter = ws1.iter_cols(min_col=min_col, max_col=max_col, min_row=min_row, max_row=max_row)
        cols_iter = list(cols_iter)
        # Start of table
        ws1.column_dimensions[cols_iter[0][1].column_letter].width = 2  # A
        # Row names
        cols_iter[1][1].border = border_t  # B
        # PASS %
        cols_iter[1][2].value = 'Pass %'  # B
        cols_iter[1][2].font = font_N  # B
        cols_iter[1][2].border = border_t  # B
        ws1.column_dimensions[cols_iter[1][2].column_letter].width = 20  # B
        # FAILS %
        cols_iter[1][3].value = '# of Failures'
        cols_iter[1][3].font = font_N
        cols_iter[1][3].border = border_t
        # EXECUTED TESTS
        cols_iter[1][4].value = '# of Tests Executed'
        cols_iter[1][4].font = font_N
        cols_iter[1][4].border = border_t
        # NOT EXECUTED TESTS
        cols_iter[1][5].value = '# of Tests Not  Executed'
        cols_iter[1][5].font = font_N
        cols_iter[1][5].border = border_t
        # END of table
        ws1.column_dimensions[cols_iter[32][1].column_letter].width = 2  # AG

        # MONTH STATS
        # Row names
        # Table header
        cols_iter[33][0].value = 'Last 30 days stats:'  # AH
        cols_iter[33][0].font = font_B  # AH
        # Last 30 days stats SUM
        ws1.column_dimensions[cols_iter[33][1].column_letter].width = 20  # AH
        cols_iter[33][1].value = '# of Days at 100%:'
        cols_iter[33][1].font = font_N
        cols_iter[33][1].border = border_t
        # Avg pass rate for 30 days
        cols_iter[33][2].value = 'Avg Pass Rate:'
        cols_iter[33][2].font = font_N
        cols_iter[33][2].border = border_t
        # Avg COUNT of failures for 30 days
        cols_iter[33][3].value = 'Avg # of Failures:'
        cols_iter[33][3].font = font_N
        cols_iter[33][3].border = border_t
        # END of table
        ws1.column_dimensions[cols_iter[35][1].column_letter].width = 2  # AG

        # FILL DATA VERTICALLY and in reverse order:
        cols_iter.reverse()
        for i, _cells in enumerate(cols_iter[4:]):
            print(f'{i} COL, _cells: {_cells} in {i} COL.')
            if i in range(0, 30):
                # print(f'{1} Col, _cells[{1}]: {_cells[1]} in {1} col. Day = i: {i}')
                test_stats = make_query(day=i)
                date = _cells[1]
                pass_rate = _cells[2]
                fail_rate = _cells[3]
                exec_rate = _cells[4]
                not_exec_rate = _cells[5]

                # date.value = 'date'
                # pass_rate.value = 'pass_rate'
                # fail_rate.value = 'fail_rate'
                # exec_rate.value = 'exec_rate'
                # not_exec_rate.value = 'not_exec_rate'

                date.value = test_stats['date'].strftime('%m/%d')  # 'date'
                if test_stats['tests_sum']:
                    pass_rate.value = f"{round(100 * (int(test_stats['passed_sum']) + int(test_stats['skipped_sum'])) / float(test_stats['tests_sum']), 1)}%"  # 'pass %'
                    fail_rate.value = test_stats['fails_sum']  # '# fails'
                    exec_rate.value = test_stats['tests_sum']  # '# executed'
                    not_exec_rate.value = 0  # '# not executed'

        wb.save(filename=place + filename)


    # insert_digest()
    # make_queries()
    # make_query(day=1)

    yearly = make_q_year()
    for q in yearly:
        print(f'Daily for the year: {q["day"].strftime("%b - %Y %d")} {q}')

    cuur_month = make_q_month()
    for q in cuur_month:
        print(f'Current month: {q["day"].strftime("%b - %Y %d")} {q}')

    days_range = make_q_days_range()
    for q in days_range:
        print(f'Days range: {q["day"].strftime("%b - %Y %d")} {q}')

    # workbook_create(filename='Discovery-Content')
    # workbook_create_test(filename='Discovery-Content')

    class GeneratorExcel:

        def __init__(self):
            if settings.DEV:
                self.place = ''
            else:
                self.place = '/home/user/TH_Octopus/UPLOAD/report/'

            self.font_N = Font(name='Calibri', size=9, bold=False)
            self.font_B = Font(name='Calibri', size=9, bold=True)

            self.border_t = Border(
                left=Side(border_style='thin', color='FF000000'),
                right=Side(border_style='thin', color='FF000000'),
                top=Side(border_style='thin', color='FF000000'),
                bottom=Side(border_style='thin', color='FF000000'),
            )
            self.border_s = Border(
                left=Side(border_style='thin', color='FF000000'),
                right=Side(border_style='thin', color='FF000000'),
            )
            self.alignment_c = Alignment(
                horizontal='center',
                vertical='center',
                shrinkToFit=True)

            self.fill_yellow = PatternFill("solid", fgColor="00FFFF99")

            # General work boor:
            self.wb = Workbook()

        def last_30_days(self):
            ws1 = self.wb.active
            ws1.title = "Last_30_days"
            # HEADER
            ws1["B1"] = 'Automation Status for BMC Discovery Content'
            ws1["B1"].font = self.font_B
            cols = self.table_scheme(ws=ws1, start_row=1, start_col=2)
            # Fill that COLs with data:
            # FILL DATA VERTICALLY and in reverse order:
            cols.reverse()
            for i, _cells in enumerate(cols[4:]):
                print(f'{i} COL, _cells: {_cells} in {i} COL.')
                # TODO: Later change to query just all 30 days?
                if i in range(0, 30):
                    # print(f'{1} Col, _cells[{1}]: {_cells[1]} in {1} col. Day = i: {i}')
                    test_stats = make_query(day=i)
                    date = _cells[1]
                    pass_rate = _cells[2]
                    fail_rate = _cells[3]
                    exec_rate = _cells[4]
                    not_exec_rate = _cells[5]

                    # date.value = 'date'
                    # pass_rate.value = 'pass_rate'
                    # fail_rate.value = 'fail_rate'
                    # exec_rate.value = 'exec_rate'
                    # not_exec_rate.value = 'not_exec_rate'

                    date.value = test_stats['date'].strftime('%m/%d')  # 'date'
                    if test_stats['tests_sum']:
                        pass_rate.value = f"{round(100 * (int(test_stats['passed_sum']) + int(test_stats['skipped_sum'])) / float(test_stats['tests_sum']), 1)}%"  # 'pass %'
                        fail_rate.value = test_stats['fails_sum']  # '# fails'
                        exec_rate.value = test_stats['tests_sum']  # '# executed'
                        not_exec_rate.value = 0  # '# not executed'

        def last_year_monthly(self):
            ws2 = self.wb.create_sheet(title="Historical")
            # Date - get all months for last year will all days.
            # Query - get last year stats - monthly (then daily) and fill the second sheet

            # TODO: Here run loop for this for each month, starting from upper row, and add + 7 on each iter
            start_row = 1
            for i in range(12):
                cols = self.table_scheme(ws=ws2, start_row=start_row, start_col=2)
                start_row += 7
                cols[1][1].value = f'{i} Date (%b)-(%Y)'  # 1st Cell of History with month and year
                cols[1][1].fill = self.fill_yellow  # B
            # Fill that COLs with data:
            # TODO: Query should be effective?

        def table_scheme(self, ws, start_row=1, start_col=2):
            end_row = start_row + 1  # 2nd row to merge with starting
            end_column = 36  # End of table in col length - AJ col in table.

            min_col = 1  # Always start iter COLs from A col - 1st
            max_col = end_column  # Always end iter on the latest COL: 36th AJ col in table.
            min_row = start_row + 2  # Where day cols are starting to count from - 3rd after two merged
            max_row = start_row + 7  # Where day stat cols and on row  - 8th from the 1st initial

            # Merging cells:
            ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=36)
            # HEAD
            ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=36)
            # HEADER
            ws["B1"] = 'Automation Status for BMC Discovery Content'
            ws["B1"].font = self.font_B

            self.horizontal(ws, min_col, max_col, min_row, max_row)
            cols = self.vertical(ws, min_col, max_col, min_row, max_row)
            return cols

        def horizontal(self, ws, min_col, max_col, min_row, max_row):
            # HORIZONTAL Iter ROW 4:8
            rows_iter = ws.iter_rows(min_col=min_col, max_col=max_col, min_row=min_row + 1, max_row=max_row)
            rows_iter = list(rows_iter)
            for i, _cells in enumerate(rows_iter):
                # print(f'{i} ROW, _cells: {_cells} in {i} ROW.')
                # ROW of 'date' - make formatting before.
                if i == 0:
                    for c in range(2, 32):
                        ws.column_dimensions[_cells[c].column_letter].width = 6
                        _cells[c].font = self.font_N
                        _cells[c].border = self.border_t
                        _cells[c].fill = self.fill_yellow
                        _cells[c].alignment = self.alignment_c
                # Row with data - make formatting before.
                if i in range(1, 4):
                    for c in range(2, 32):
                        ws.column_dimensions[_cells[c].column_letter].width = 6
                        _cells[c].font = self.font_N
                        _cells[c].border = self.border_s
                        _cells[c].alignment = self.alignment_c
                # ROW to put test report data - make formatting before.
                if i in range(0, 4):
                    for c in range(2, 32):
                        ws.column_dimensions[_cells[c].column_letter].width = 6
                        # _cells[c].value = '0'
                        _cells[c].font = self.font_N
                        # _cells[c].border = self.border_s
                        _cells[c].alignment = self.alignment_c
                # ROW of '#not executed' - make formatting before and fill 0 by default, as we run 100% of tests!
                if i == 4:
                    for c in range(2, 32):
                        ws.column_dimensions[_cells[c].column_letter].width = 6
                        _cells[c].value = '0'  # Do not change, we don't actually  "NOT RUN" the tests
                        _cells[c].font = self.font_N
                        _cells[c].border = self.border_t
                        _cells[c].alignment = self.alignment_c

        def vertical(self, ws, min_col, max_col, min_row, max_row):
            # VERTICAL Iter COL A:AJ
            cols_iter = ws.iter_cols(min_col=min_col, max_col=max_col, min_row=min_row, max_row=max_row)
            cols_iter = list(cols_iter)
            # Start of table
            ws.column_dimensions[cols_iter[0][1].column_letter].width = 2  # A
            # Row names
            cols_iter[1][1].border = self.border_t  # B
            # PASS %
            cols_iter[1][2].value = 'Pass %'  # B
            cols_iter[1][2].font = self.font_N  # B
            cols_iter[1][2].border = self.border_t  # B
            ws.column_dimensions[cols_iter[1][2].column_letter].width = 20  # B
            # FAILS %
            cols_iter[1][3].value = '# of Failures'
            cols_iter[1][3].font = self.font_N
            cols_iter[1][3].border = self.border_t
            # EXECUTED TESTS
            cols_iter[1][4].value = '# of Tests Executed'
            cols_iter[1][4].font = self.font_N
            cols_iter[1][4].border = self.border_t
            # NOT EXECUTED TESTS
            cols_iter[1][5].value = '# of Tests Not  Executed'
            cols_iter[1][5].font = self.font_N
            cols_iter[1][5].border = self.border_t
            # END of table
            ws.column_dimensions[cols_iter[32][1].column_letter].width = 2  # AG

            # MONTH STATS
            # Row names
            # Table header
            cols_iter[33][0].value = 'Last 30 days stats:'  # AH
            cols_iter[33][0].font = self.font_B  # AH
            # Last 30 days stats SUM
            ws.column_dimensions[cols_iter[33][1].column_letter].width = 20  # AH
            cols_iter[33][1].value = '# of Days at 100%:'
            cols_iter[33][1].font = self.font_N
            cols_iter[33][1].border = self.border_t
            # Avg pass rate for 30 days
            cols_iter[33][2].value = 'Avg Pass Rate:'
            cols_iter[33][2].font = self.font_N
            cols_iter[33][2].border = self.border_t
            # Avg COUNT of failures for 30 days
            cols_iter[33][3].value = 'Avg # of Failures:'
            cols_iter[33][3].font = self.font_N
            cols_iter[33][3].border = self.border_t
            # END of table
            ws.column_dimensions[cols_iter[35][1].column_letter].width = 2  # AG
            # For further fill data in cols
            return cols_iter

        def main_report(self, filename):
            self.last_30_days()
            self.last_year_monthly()
            self.wb.save(filename=self.place + filename + '.xlsx')

    GeneratorExcel().main_report(filename='test_wb_classed')
