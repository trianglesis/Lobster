import datetime
import logging

from django.db.models import Q
from django.db.models.aggregates import Sum, Max
from django.db.utils import IntegrityError

from octo_tku_patterns.model_views import TestReportsView
from octo_tku_patterns.models import TestReports

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Fill

log = logging.getLogger("octo.octologger")


class Report:

    @staticmethod
    def get_stats(**kwargs):
        # print(f"Selecting by: {kwargs}")
        report_date_time = kwargs.get('report_date_time')

        if not isinstance(report_date_time, datetime.date):
            print(f"Warning: report_date_time should bw a datetime object, if not - will use today date by default!")
            date_var = datetime.date.today()
        else:
            date_var = report_date_time

        queryset = TestReports.objects.all()
        if kwargs.get("test_type"):
            queryset = queryset.filter(
                test_type__exact=kwargs.get("test_type")
            )
        else:
            queryset = queryset.filter(
                tkn_branch__isnull=False)
        if kwargs.get('addm_name'):
            queryset = queryset.filter(
                addm_name__exact=kwargs.get('addm_name')
            )
        if kwargs.get('tkn_branch'):
            queryset = queryset.filter(
                tkn_branch__exact=kwargs.get('tkn_branch')
            )
        if kwargs.get('pattern_library'):
            queryset = queryset.filter(
                pattern_library__exact=kwargs.get('pattern_library')
            )
        if kwargs.get('report_date_time'):
            queryset = queryset.filter(
                Q(report_date_time__year=date_var.year,
                  report_date_time__month=date_var.month,
                  report_date_time__day=date_var.day)
            )

        return queryset

    @staticmethod
    def insert_digest():
        tests = TestReportsView.objects.all().values(
            # 'test_type',
            # 'tkn_branch',
            # 'pattern_library',
            # 'addm_name',
            # 'addm_v_int',
            # 'tests_count',
            # 'patterns_count',
            # 'fails',
            # 'error',
            # 'passed',
            # 'skipped',
        )
        for item in tests:
            try:
                save_tst = TestReports(**item)
                save_tst.save(force_insert=True)
            except IntegrityError as e:
                print(f'Duplicate: {e}')

    def select_stats(self, **kwargs):
        by_value = kwargs.get('grouping')
        queryset = self.get_stats(**kwargs)
        queryset = queryset.values(by_value).order_by(by_value).annotate(
            date=Max('report_date_time'),
            addm_v_int=Max('addm_v_int'),
            tests_sum=Sum('tests_count'),
            patterns_sum=Sum('patterns_count'),
            errors_sum=Sum('error'),
            fails_sum=Sum('fails'),
            passed_sum=Sum('passed'),
            skipped_sum=Sum('skipped'),
        )
        # print(f"{queryset.query}")
        return queryset

    def generate_xlxs(self, **kwargs):
        log.debug(f"Generating XML dummy! {kwargs}")


class GeneratorExcel:

    def __init__(self):
        self.place = '/home/user/TH_Octopus/UPLOAD/report/'

        self.font_N = Font(name='Calibri', size=9, bold=False)
        self.font_B = Font(name='Calibri', size=9, bold=True)

        self.border_t = Border(
            left=Side(border_style='thin', color='FF000000'),
            right=Side(border_style='thin', color='FF000000'),
            top=Side(border_style='thin', color='FF000000'),
            bottom=Side(border_style='thin', color='FF000000'),
        )
        self.border_s = Border(
            left=Side(border_style='thin', color='FF000000'),
            right=Side(border_style='thin', color='FF000000'),
        )
        self.alignment_c = Alignment(
            horizontal='center',
            vertical='center',
            shrinkToFit=True)

        self.fill_yellow = PatternFill("solid", fgColor="00FFFF99")

        # General work boor:
        self.wb = Workbook()

    def last_30_days(self):
        ws1 = self.wb.active
        ws1.title = "Last_30_days"
        # HEADER
        ws1["B1"] = 'Automation Status for BMC Discovery Content'
        ws1["B1"].font = self.font_B
        self.table_scheme(ws=ws1, start_row=1, start_col=2)

    def last_year_monthly(self):
        ws2 = self.wb.create_sheet(title="Historical")
        # Date - get all months for last year will all days.
        # Query - get last year stats - monthly (then daily) and fill the second sheet

        # TODO: Here run loop for this for each month, starting from upper row, and add + 7 on each iter
        self.table_scheme(ws=ws2, start_row=1, start_col=2)

    def table_scheme(self, ws, start_row=1, start_col=2):
        end_row = start_row + 1  # 2nd row to merge with starting
        end_column = 36          # End of table in col length - AJ col in table.

        min_col = 1              # Always start iter COLs from A col - 1st
        max_col = end_column     # Always end iter on the latest COL: 36th AJ col in table.
        min_row = start_row + 2  # Where day cols are starting to count from - 3rd after two merged
        max_row = start_row + 7  # Where day stat cols and on row  - 8th from the 1st initial

        # Merging cells:
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_column)
        # HEAD
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_column)
        # HEADER
        ws["B1"] = 'Automation Status for BMC Discovery Content'
        ws["B1"].font = self.font_B

        self.horizontal(ws, min_col, max_col, min_row, max_row)
        self.vertical(ws, min_col, max_col, min_row, max_row)

    def horizontal(self, ws, min_col, max_col, min_row, max_row):
        # HORIZONTAL Iter ROW 4:8
        rows_iter = ws.iter_rows(min_col=min_col, max_col=max_col, min_row=min_row+1, max_row=max_row)
        rows_iter = list(rows_iter)
        for i, _cells in enumerate(rows_iter):
            # print(f'{i} ROW, _cells: {_cells} in {i} ROW.')
            # ROW of 'date' - make formatting before.
            if i == 0:
                for c in range(2, 32):
                    ws.column_dimensions[_cells[c].column_letter].width = 6
                    _cells[c].font = self.font_N
                    _cells[c].border = self.border_t
                    _cells[c].fill = self.fill_yellow
                    _cells[c].alignment = self.alignment_c
            # Row with data - make formatting before.
            if i in range(1, 4):
                for c in range(2, 32):
                    ws.column_dimensions[_cells[c].column_letter].width = 6
                    _cells[c].font = self.font_N
                    _cells[c].border = self.border_s
                    _cells[c].alignment = self.alignment_c
            # ROW to put test report data - make formatting before.
            if i in range(0, 4):
                for c in range(2, 32):
                    ws.column_dimensions[_cells[c].column_letter].width = 6
                    # _cells[c].value = '0'
                    _cells[c].font = self.font_N
                    # _cells[c].border = self.border_s
                    _cells[c].alignment = self.alignment_c
            # ROW of '#not executed' - make formatting before and fill 0 by default, as we run 100% of tests!
            if i == 4:
                for c in range(2, 32):
                    ws.column_dimensions[_cells[c].column_letter].width = 6
                    _cells[c].value = '0'  # Do not change, we don't actually  "NOT RUN" the tests
                    _cells[c].font = self.font_N
                    _cells[c].border = self.border_t
                    _cells[c].alignment = self.alignment_c

    def vertical(self, ws, min_col, max_col, min_row, max_row):
        # VERTICAL Iter COL A:AJ
        cols_iter = ws.iter_cols(min_col=min_col, max_col=max_col, min_row=min_row, max_row=max_row)
        cols_iter = list(cols_iter)
        # Start of table
        ws.column_dimensions[cols_iter[0][1].column_letter].width = 2  # A
        # Row names
        cols_iter[1][1].border = self.border_t  # B
        # PASS %
        cols_iter[1][2].value = 'Pass %'  # B
        cols_iter[1][2].font = self.font_N  # B
        cols_iter[1][2].border = self.border_t  # B
        ws.column_dimensions[cols_iter[1][2].column_letter].width = 20  # B
        # FAILS %
        cols_iter[1][3].value = '# of Failures'
        cols_iter[1][3].font = self.font_N
        cols_iter[1][3].border = self.border_t
        # EXECUTED TESTS
        cols_iter[1][4].value = '# of Tests Executed'
        cols_iter[1][4].font = self.font_N
        cols_iter[1][4].border = self.border_t
        # NOT EXECUTED TESTS
        cols_iter[1][5].value = '# of Tests Not  Executed'
        cols_iter[1][5].font = self.font_N
        cols_iter[1][5].border = self.border_t
        # END of table
        ws.column_dimensions[cols_iter[32][1].column_letter].width = 2  # AG

        # MONTH STATS
        # Row names
        # Table header
        cols_iter[33][0].value = 'Last 30 days stats:'  # AH
        cols_iter[33][0].font = self.font_B  # AH
        # Last 30 days stats SUM
        ws.column_dimensions[cols_iter[33][1].column_letter].width = 20  # AH
        cols_iter[33][1].value = '# of Days at 100%:'
        cols_iter[33][1].font = self.font_N
        cols_iter[33][1].border = self.border_t
        # Avg pass rate for 30 days
        cols_iter[33][2].value = 'Avg Pass Rate:'
        cols_iter[33][2].font = self.font_N
        cols_iter[33][2].border = self.border_t
        # Avg COUNT of failures for 30 days
        cols_iter[33][3].value = 'Avg # of Failures:'
        cols_iter[33][3].font = self.font_N
        cols_iter[33][3].border = self.border_t
        # END of table
        ws.column_dimensions[cols_iter[35][1].column_letter].width = 2  # AG

    def main_report(self, filename):
        self.last_30_days()
        self.last_year_monthly()
        self.wb.save(filename=self.place + filename + '.xlsx')
